#!/usr/bin/env python3
"""
AI 스마트 가계부 - CSV 및 Excel 파일 파싱
Author: leehansol
Created: 2025-05-25
"""

import csv
import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import chardet


class FileParser:
    """CSV 및 Excel 파일 파싱을 담당하는 클래스"""
    
    @staticmethod
    def parse_csv_preview(file_path: str, max_rows: int = 5) -> Dict:
        """
        CSV 파일의 첫 N행을 파싱하여 미리보기 데이터 반환
        
        Args:
            file_path: CSV 파일 경로
            max_rows: 추출할 최대 행 수 (헤더 제외)
            
        Returns:
            dict: 파싱 결과 정보
                - success: 파싱 성공 여부
                - headers: 헤더 행 리스트
                - data: 데이터 행 리스트
                - total_rows: 총 데이터 행 수 (헤더 제외)
                - error: 오류 메시지 (실패 시)
        """
        result = {
            'success': False,
            'headers': [],
            'data': [],
            'total_rows': 0,
            'error': None
        }
        
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                result['error'] = f"파일이 존재하지 않습니다: {file_path}"
                return result
            
            # 파일 크기 확인 (빈 파일 체크)
            if os.path.getsize(file_path) == 0:
                result['error'] = "파일이 비어있습니다"
                return result
            
            # 인코딩 감지 시도
            encoding = 'utf-8'
            try:
                with open(file_path, 'rb') as f:
                    raw_data = f.read(1024)
                    if raw_data:
                        detected = chardet.detect(raw_data)
                        if detected['encoding'] and detected['confidence'] > 0.7:
                            encoding = detected['encoding']
            except:
                # 인코딩 감지 실패 시 UTF-8 사용
                pass
            
            # CSV 파일 읽기
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as csvfile:
                    # CSV 방언 자동 감지 시도
                    try:
                        sample = csvfile.read(1024)
                        csvfile.seek(0)
                        sniffer = csv.Sniffer()
                        dialect = sniffer.sniff(sample)
                    except:
                        # 감지 실패 시 기본 설정 사용
                        dialect = csv.excel
                    
                    # CSV 리더 생성
                    reader = csv.reader(csvfile, dialect)
                    
                    # 헤더 행 읽기
                    try:
                        headers = next(reader)
                        # 헤더 유효성 검증
                        if not headers or all(not str(h).strip() for h in headers):
                            result['error'] = "유효한 헤더가 없습니다. 첫 번째 행이 비어있습니다."
                            return result
                        
                        result['headers'] = headers
                        print(f"📋 헤더 발견: {headers}")
                    except StopIteration:
                        result['error'] = "파일이 비어있습니다"
                        return result
                    
                    # 데이터 행 읽기 (최대 max_rows개)
                    data_rows = []
                    total_count = 0
                    malformed_rows = 0
                    
                    for row_num, row in enumerate(reader, start=2):  # 헤더 다음부터 시작
                        total_count += 1
                        
                        # 행 품질 검증
                        if len(row) != len(headers):
                            malformed_rows += 1
                            # 경고는 하지만 계속 진행
                            if malformed_rows <= 3:  # 처음 3개 오류만 로깅
                                print(f"⚠️ {row_num}행: 컬럼 수 불일치 (헤더: {len(headers)}, 데이터: {len(row)})")
                        
                        if len(data_rows) < max_rows:
                            data_rows.append(row)
                    
                    result['data'] = data_rows
                    result['total_rows'] = total_count
                    result['success'] = True
                    
                    # 품질 경고
                    if malformed_rows > 0:
                        print(f"⚠️ 주의: {malformed_rows}개 행에서 컬럼 수 불일치가 발견되었습니다.")
                    
                    print(f"📊 데이터 행 {len(data_rows)}개 추출 (전체 {total_count}개 중)")
                    
            except UnicodeDecodeError as e:
                # UTF-8로 실패한 경우 다른 인코딩 시도
                fallback_encodings = ['cp949', 'euc-kr', 'latin-1']
                for fallback_encoding in fallback_encodings:
                    try:
                        with open(file_path, 'r', encoding=fallback_encoding, newline='') as csvfile:
                            # 간단한 테스트 읽기
                            csvfile.read(100)
                            result['error'] = f"인코딩 오류: 파일이 {encoding} 형식이 아닙니다. {fallback_encoding} 인코딩을 시도해보세요."
                            return result
                    except:
                        continue
                result['error'] = f"인코딩 오류: {e}. 파일이 UTF-8 형식이 아니며 자동 감지에 실패했습니다."
                
        except csv.Error as e:
            result['error'] = f"CSV 형식 오류: {e}. 파일이 올바른 CSV 형식이 아닙니다."
        except PermissionError:
            result['error'] = "파일 접근 권한이 없습니다."
        except Exception as e:
            result['error'] = f"파일 읽기 오류: {e}"
        
        return result
    
    @staticmethod
    def parse_excel_preview(file_path: str, max_rows: int = 5) -> Dict:
        """
        Excel 파일의 첫 N행을 파싱하여 미리보기 데이터 반환
        
        Args:
            file_path: Excel 파일 경로 (XLS, XLSX)
            max_rows: 추출할 최대 행 수 (헤더 제외)
            
        Returns:
            dict: 파싱 결과 정보
                - success: 파싱 성공 여부
                - headers: 헤더 행 리스트
                - data: 데이터 행 리스트
                - total_rows: 총 데이터 행 수 (헤더 제외)
                - error: 오류 메시지 (실패 시)
        """
        result = {
            'success': False,
            'headers': [],
            'data': [],
            'total_rows': 0,
            'error': None
        }
        
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                result['error'] = f"파일이 존재하지 않습니다: {file_path}"
                return result
            
            # 파일 크기 확인
            if os.path.getsize(file_path) == 0:
                result['error'] = "파일이 비어있습니다"
                return result
            
            # Excel 파일 읽기
            try:
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            except InvalidFileException as e:
                result['error'] = f"Excel 파일 형식 오류: {e}. 올바른 Excel 파일이 아닙니다."
                return result
            except PermissionError:
                result['error'] = "파일 접근 권한이 없습니다."
                return result
            
            # 첫 번째 시트 선택
            worksheet = workbook.active
            
            # 워크시트가 비어있는지 확인
            if worksheet.max_row is None or worksheet.max_row == 0:
                workbook.close()
                result['error'] = "파일이 비어있습니다"
                return result
            
            # 모든 행을 리스트로 변환
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()
            
            # 빈 행들 제거
            non_empty_rows = []
            for row in rows:
                # 모든 셀이 None이 아닌 행만 포함
                if any(cell is not None for cell in row):
                    # None 값들을 빈 문자열로 변환하고, 숫자를 문자열로 변환
                    converted_row = []
                    for cell in row:
                        if cell is None:
                            converted_row.append('')
                        else:
                            converted_row.append(str(cell))
                    non_empty_rows.append(converted_row)
            
            if not non_empty_rows:
                result['error'] = "파일이 비어있습니다"
                return result
            
            # 헤더 행 추출 (첫 번째 행)
            headers = non_empty_rows[0]
            
            # 헤더 유효성 검증
            if not headers or all(not str(h).strip() for h in headers):
                result['error'] = "유효한 헤더가 없습니다. 첫 번째 행이 비어있습니다."
                return result
            
            result['headers'] = headers
            print(f"📋 헤더 발견: {headers}")
            
            # 데이터 행 추출 (헤더 이후 행들)
            data_rows = non_empty_rows[1:]
            total_count = len(data_rows)
            
            # 데이터 품질 검증
            malformed_rows = 0
            for row_num, row in enumerate(data_rows, start=2):
                if len(row) != len(headers):
                    malformed_rows += 1
                    if malformed_rows <= 3:  # 처음 3개 오류만 로깅
                        print(f"⚠️ {row_num}행: 컬럼 수 불일치 (헤더: {len(headers)}, 데이터: {len(row)})")
            
            # 요청된 수만큼만 반환
            preview_data = data_rows[:max_rows]
            
            result['data'] = preview_data
            result['total_rows'] = total_count
            result['success'] = True
            
            # 품질 경고
            if malformed_rows > 0:
                print(f"⚠️ 주의: {malformed_rows}개 행에서 컬럼 수 불일치가 발견되었습니다.")
            
            print(f"📊 데이터 행 {len(preview_data)}개 추출 (전체 {total_count}개 중)")
                
        except Exception as e:
            result['error'] = f"Excel 파일 읽기 오류: {e}"
        
        return result
    
    @staticmethod
    def print_csv_preview(parse_result: Dict) -> None:
        """
        파싱 결과를 콘솔에 예쁘게 출력
        
        Args:
            parse_result: parse_csv_preview 함수의 반환값
        """
        if not parse_result['success']:
            print(f"❌ CSV 파싱 실패: {parse_result['error']}")
            return
        
        headers = parse_result['headers']
        data = parse_result['data']
        total_rows = parse_result['total_rows']
        
        print("\n" + "="*60)
        print("📄 CSV 파일 미리보기")
        print("="*60)
        
        # 헤더 출력
        print(f"📋 헤더 ({len(headers)}개 컬럼):")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. {header}")
        
        print(f"\n📊 데이터 미리보기 (첫 {len(data)}행 / 전체 {total_rows}행):")
        
        if not data:
            print("  (데이터 없음)")
        else:
            # 각 데이터 행 출력
            for i, row in enumerate(data, 1):
                print(f"\n  📝 {i}행:")
                for j, (header, value) in enumerate(zip(headers, row)):
                    print(f"    {header}: {value}")
        
        print("="*60 + "\n")
    
    @staticmethod
    def get_file_summary(file_path: str) -> Dict:
        """
        CSV 파일의 기본 정보 반환 (전체 스캔)
        
        Args:
            file_path: CSV 파일 경로
            
        Returns:
            dict: 파일 요약 정보
        """
        try:
            with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                reader = csv.reader(csvfile)
                
                # 헤더 읽기
                headers = next(reader)
                
                # 전체 행 수 계산
                row_count = sum(1 for row in reader)
                
                return {
                    'success': True,
                    'column_count': len(headers),
                    'row_count': row_count,
                    'headers': headers
                }
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }


# 편의 함수들
def parse_csv_file(file_path: str, max_rows: int = 5) -> Dict:
    """FileParser.parse_csv_preview의 편의 함수"""
    return FileParser.parse_csv_preview(file_path, max_rows)


def parse_excel_file(file_path: str, max_rows: int = 5) -> Dict:
    """FileParser.parse_excel_preview의 편의 함수"""
    return FileParser.parse_excel_preview(file_path, max_rows)


def print_csv_file(file_path: str, max_rows: int = 5) -> None:
    """CSV 파일을 파싱하고 바로 콘솔에 출력하는 편의 함수"""
    result = FileParser.parse_csv_preview(file_path, max_rows)
    FileParser.print_csv_preview(result)


def print_excel_file(file_path: str, max_rows: int = 5) -> None:
    """Excel 파일을 파싱하고 바로 콘솔에 출력하는 편의 함수"""
    result = FileParser.parse_excel_preview(file_path, max_rows)
    FileParser.print_csv_preview(result) 