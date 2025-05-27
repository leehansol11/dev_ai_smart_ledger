#!/usr/bin/env python3
"""
AI 스마트 가계부 - CSV 파일 파싱 테스트
Author: leehansol
Created: 2025-05-25
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from ai_smart_ledger.app.core.file_parser import FileParser, parse_csv_file, print_csv_file, parse_excel_file, print_excel_file


class TestFileParser:
    """FileParser 클래스 테스트"""
    
    @pytest.fixture
    def temp_csv_with_header(self):
        """헤더가 있는 테스트용 CSV 파일 생성"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
            f.write("날짜,내용,금액,카테고리\n")
            f.write("2023-01-01,스타벅스 커피,4500,식비\n")
            f.write("2023-01-02,버스 요금,1350,교통비\n")
            f.write("2023-01-03,급여 입금,2500000,수입\n")
            f.write("2023-01-04,마트 장보기,35000,식비\n")
            f.write("2023-01-05,영화 관람,15000,여가\n")
            f.write("2023-01-06,휴대폰 요금,55000,통신비\n")
            f.write("2023-01-07,점심 식사,8000,식비\n")
            temp_path = f.name
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_csv_no_header(self):
        """헤더가 없는 테스트용 CSV 파일 생성"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
            f.write("2023-01-01,스타벅스 커피,4500\n")
            f.write("2023-01-02,버스 요금,1350\n")
            f.write("2023-01-03,급여 입금,2500000\n")
            temp_path = f.name
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_csv_minimal(self):
        """최소 데이터(2행)를 가진 CSV 파일 생성"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
            f.write("날짜,내용\n")
            f.write("2023-01-01,테스트\n")
            temp_path = f.name
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_csv_empty(self):
        """비어있는 CSV 파일 생성"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
            # 빈 파일
            temp_path = f.name
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_excel_with_header(self):
        """헤더가 있는 테스트용 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        
        # 헤더 행 추가
        ws.append(["날짜", "내용", "금액", "카테고리"])
        
        # 데이터 행 추가
        ws.append(["2023-01-01", "스타벅스 커피", 4500, "식비"])
        ws.append(["2023-01-02", "버스 요금", 1350, "교통비"])
        ws.append(["2023-01-03", "급여 입금", 2500000, "수입"])
        ws.append(["2023-01-04", "마트 장보기", 35000, "식비"])
        ws.append(["2023-01-05", "영화 관람", 15000, "여가"])
        ws.append(["2023-01-06", "휴대폰 요금", 55000, "통신비"])
        ws.append(["2023-01-07", "점심 식사", 8000, "식비"])
        
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        wb.save(temp_path)
        wb.close()
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_excel_minimal(self):
        """최소 데이터(2행)를 가진 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["날짜", "내용"])
        ws.append(["2023-01-01", "테스트"])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        wb.save(temp_path)
        wb.close()
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_excel_empty(self):
        """빈 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        # 데이터 없이 빈 워크시트
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        wb.save(temp_path)
        wb.close()
        
        yield temp_path
        
        # 정리
        if os.path.exists(temp_path):
            os.unlink(temp_path)

    def test_file_parser_class_exists(self):
        """FileParser 클래스 존재 테스트"""
        assert FileParser is not None
        assert hasattr(FileParser, 'parse_csv_preview')
        assert hasattr(FileParser, 'print_csv_preview')
        assert hasattr(FileParser, 'get_file_summary')
        # Excel 파싱 메서드 존재 확인
        assert hasattr(FileParser, 'parse_excel_preview')
    
    def test_parse_csv_with_header_success(self, temp_csv_with_header):
        """헤더가 있는 CSV 파일 파싱 성공 테스트"""
        result = FileParser.parse_csv_preview(temp_csv_with_header, max_rows=5)
        
        assert result['success'] is True
        assert result['error'] is None
        assert len(result['headers']) == 4
        assert result['headers'] == ['날짜', '내용', '금액', '카테고리']
        assert len(result['data']) == 5  # 첫 5행
        assert result['total_rows'] == 7  # 전체 데이터 행 수
        
        # 첫 번째 데이터 행 확인
        first_row = result['data'][0]
        assert first_row == ['2023-01-01', '스타벅스 커피', '4500', '식비']
    
    def test_parse_csv_with_header_limited_rows(self, temp_csv_with_header):
        """제한된 행 수로 CSV 파싱 테스트"""
        result = FileParser.parse_csv_preview(temp_csv_with_header, max_rows=3)
        
        assert result['success'] is True
        assert len(result['data']) == 3  # 요청한 3행만
        assert result['total_rows'] == 7  # 전체는 7행
    
    def test_parse_csv_no_header(self, temp_csv_no_header):
        """헤더가 없는 CSV 파일 파싱 테스트"""
        result = FileParser.parse_csv_preview(temp_csv_no_header, max_rows=5)
        
        assert result['success'] is True
        # 첫 번째 행이 헤더로 인식됨
        assert result['headers'] == ['2023-01-01', '스타벅스 커피', '4500']
        assert len(result['data']) == 2  # 나머지 2행
        assert result['total_rows'] == 2
    
    def test_parse_csv_minimal_data(self, temp_csv_minimal):
        """최소 데이터 CSV 파일 파싱 테스트"""
        result = FileParser.parse_csv_preview(temp_csv_minimal, max_rows=5)
        
        assert result['success'] is True
        assert result['headers'] == ['날짜', '내용']
        assert len(result['data']) == 1
        assert result['total_rows'] == 1
        assert result['data'][0] == ['2023-01-01', '테스트']
    
    def test_parse_csv_empty_file(self, temp_csv_empty):
        """비어있는 CSV 파일 파싱 테스트"""
        result = FileParser.parse_csv_preview(temp_csv_empty, max_rows=5)
        
        assert result['success'] is False
        assert result['error'] == "파일이 비어있습니다"
        assert len(result['headers']) == 0
        assert len(result['data']) == 0
    
    def test_parse_csv_nonexistent_file(self):
        """존재하지 않는 파일 파싱 테스트"""
        result = FileParser.parse_csv_preview("/nonexistent/file.csv", max_rows=5)
        
        assert result['success'] is False
        assert "파일이 존재하지 않습니다" in result['error']
        assert len(result['headers']) == 0
        assert len(result['data']) == 0
    
    def test_print_csv_preview_success(self, temp_csv_with_header, capsys):
        """CSV 미리보기 출력 테스트 (성공)"""
        result = FileParser.parse_csv_preview(temp_csv_with_header, max_rows=2)
        FileParser.print_csv_preview(result)
        
        captured = capsys.readouterr()
        assert "CSV 파일 미리보기" in captured.out
        assert "헤더 (4개 컬럼)" in captured.out
        assert "데이터 미리보기 (첫 2행 / 전체 7행)" in captured.out
        assert "스타벅스 커피" in captured.out
    
    def test_print_csv_preview_failure(self, capsys):
        """CSV 미리보기 출력 테스트 (실패)"""
        result = {
            'success': False,
            'error': '테스트 오류',
            'headers': [],
            'data': [],
            'total_rows': 0
        }
        FileParser.print_csv_preview(result)
        
        captured = capsys.readouterr()
        assert "CSV 파싱 실패: 테스트 오류" in captured.out
    
    def test_get_file_summary_success(self, temp_csv_with_header):
        """파일 요약 정보 조회 성공 테스트"""
        summary = FileParser.get_file_summary(temp_csv_with_header)
        
        assert summary['success'] is True
        assert summary['column_count'] == 4
        assert summary['row_count'] == 7
        assert summary['headers'] == ['날짜', '내용', '금액', '카테고리']
    
    def test_get_file_summary_failure(self):
        """파일 요약 정보 조회 실패 테스트"""
        summary = FileParser.get_file_summary("/nonexistent/file.csv")
        
        assert summary['success'] is False
        assert 'error' in summary
    
    def test_convenience_functions(self, temp_csv_with_header):
        """편의 함수들 테스트"""
        # parse_csv_file 편의 함수 테스트
        result = parse_csv_file(temp_csv_with_header, max_rows=3)
        assert result['success'] is True
        assert len(result['data']) == 3
        
        # print_csv_file 편의 함수 테스트 (출력만 확인)
        try:
            print_csv_file(temp_csv_with_header, max_rows=2)
            # 예외 없이 실행되면 성공
            assert True
        except Exception:
            pytest.fail("print_csv_file 편의 함수 실행 중 오류 발생")

    def test_parse_excel_with_header_success(self, temp_excel_with_header):
        """헤더가 있는 Excel 파일 파싱 성공 테스트"""
        result = FileParser.parse_excel_preview(temp_excel_with_header, max_rows=5)
        
        assert result['success'] is True
        assert result['error'] is None
        assert len(result['headers']) == 4
        assert result['headers'] == ['날짜', '내용', '금액', '카테고리']
        assert len(result['data']) == 5  # 첫 5행
        assert result['total_rows'] == 7  # 전체 데이터 행 수
        
        # 첫 번째 데이터 행 확인 (숫자는 문자열로 변환되어야 함)
        first_row = result['data'][0]
        assert first_row == ['2023-01-01', '스타벅스 커피', '4500', '식비']
    
    def test_parse_excel_with_header_limited_rows(self, temp_excel_with_header):
        """제한된 행 수로 Excel 파싱 테스트"""
        result = FileParser.parse_excel_preview(temp_excel_with_header, max_rows=3)
        
        assert result['success'] is True
        assert len(result['data']) == 3  # 요청한 3행만
        assert result['total_rows'] == 7  # 전체는 7행
    
    def test_parse_excel_minimal_data(self, temp_excel_minimal):
        """최소 데이터 Excel 파일 파싱 테스트"""
        result = FileParser.parse_excel_preview(temp_excel_minimal, max_rows=5)
        
        assert result['success'] is True
        assert result['headers'] == ['날짜', '내용']
        assert len(result['data']) == 1
        assert result['total_rows'] == 1
        assert result['data'][0] == ['2023-01-01', '테스트']
    
    def test_parse_excel_empty_file(self, temp_excel_empty):
        """빈 Excel 파일 파싱 테스트"""
        result = FileParser.parse_excel_preview(temp_excel_empty, max_rows=5)
        
        assert result['success'] is False
        assert result['error'] == "파일이 비어있습니다"
        assert len(result['headers']) == 0
        assert len(result['data']) == 0
    
    def test_parse_excel_nonexistent_file(self):
        """존재하지 않는 Excel 파일 파싱 테스트"""
        result = FileParser.parse_excel_preview("/nonexistent/file.xlsx", max_rows=5)
        
        assert result['success'] is False
        assert "파일이 존재하지 않습니다" in result['error']
        assert len(result['headers']) == 0
        assert len(result['data']) == 0

    def test_excel_convenience_functions(self, temp_excel_with_header):
        """Excel 편의 함수들 테스트"""
        # parse_excel_file 편의 함수 테스트
        result = parse_excel_file(temp_excel_with_header, max_rows=3)
        assert result['success'] is True
        assert len(result['data']) == 3
        
        # print_excel_file 편의 함수 테스트 (출력만 확인)
        try:
            print_excel_file(temp_excel_with_header, max_rows=2)
            # 예외 없이 실행되면 성공
            assert True
        except Exception:
            pytest.fail("print_excel_file 편의 함수 실행 중 오류 발생")


if __name__ == "__main__":
    pytest.main([__file__, "-v"]) 